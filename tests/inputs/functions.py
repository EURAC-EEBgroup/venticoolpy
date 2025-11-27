"""
    Functions to get the test inputs data from VCdesign Excel files. 
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval

from venticoolpy.model import Building, ClimateData


def get_appartment_bld_climate_data() -> ClimateData:
    outdoor_dry_bulb_temperature = pd.read_csv("tests/inputs/data/outdor_dry_bulb_temperature.csv")
    relative_humidity_outdoor_air = pd.read_csv("tests/inputs/data/outdor_climate_data.csv")
    isol_tot = pd.read_csv("tests/inputs/data/isol_tot.csv")

    climate_data = ClimateData(
        df_outdoor_dry_bulb_temperature=outdoor_dry_bulb_temperature,
        df_relative_humidity_outdoor_air=relative_humidity_outdoor_air,
        df_isol_tot=isol_tot
    )
    return climate_data



def load_climate_data_from_VCdesign(filename):

    def load_workbook_range(range_string, ws):
        col_start, col_end = re.findall("[A-Z]+", range_string)

        data_rows = []
        for row in ws[range_string]:
            data_rows.append([cell.value for cell in row])

        return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))


    wb = load_workbook(filename, data_only=True)
    # Load Inputs
    sh = wb["WEATHER FILE"]
    outdoor_dry_bulb_temperature = load_workbook_range('D20:D8779', sh)
    relative_humidity_outdoor_air = load_workbook_range('F20:F8779', sh)
    isol_tot = load_workbook_range('K20:K8779', sh)

    climate_data = ClimateData(
        df_outdoor_dry_bulb_temperature=outdoor_dry_bulb_temperature,
        df_relative_humidity_outdoor_air=relative_humidity_outdoor_air,
        df_isol_tot=isol_tot,
    )

    return climate_data